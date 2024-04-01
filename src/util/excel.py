from openpyxl import load_workbook
from loguru import logger
from .sqlite import SQLite
import pandas as pd
from datetime import datetime
from .comparison_util import compare_count
import re


def verify_excel_header(sheet_name, verify_header_df, original_data_stream):
    header = int(verify_header_df["表头在第几行"].iloc[0]) - 1
    column_start = verify_header_df["开始列"].iloc[0]
    column_end = verify_header_df["结束列"].iloc[0]
    column_end_index = transform_excel_column_to_index(column_end)
    fetch_column_start = verify_header_df["取值开始列"].iloc[0]
    fetch_column_end = verify_header_df["取值结束列"].iloc[0]
    table_name = verify_header_df["表名"].iloc[0]
    # Get oss_check_file table columns
    x = transform_excel_column_to_index(fetch_column_start)
    y = transform_excel_column_to_index(fetch_column_end)

    verify_column_list = (
        verify_header_df.iloc[:, x - 1 : y]
        .iloc[0]
        .fillna("")
        .str.replace(" ", "")
        .str.replace("\t", "")
        .str.replace("\n", "")
        .tolist()
    )

    original_sheet_name_list = [sheet_name]
    original_data_file_df, original_sheet_name_list = get_excel_df(
        data_stream=original_data_stream,
        custom_sheet_list=original_sheet_name_list,
        header=header,
        X=column_start,
        Y=column_end,
    )
    table_data_df = original_data_file_df[sheet_name]

    # Handle field renaming due to duplicate column names
    formatted_column_list_tmp = [
        col.split(".1")[0].split(".2")[0].split(".3")[0]
        if ".1" or ".2" or ".3" in col
        else col
        for col in table_data_df.columns.tolist()
    ]

    # Format columns
    formatted_column_list = [
        re.sub(r"^Unnamed:\s*\d+", "", item)
        .replace(" ", "")
        .replace("\t", "")
        .replace("\n", "")
        for item in formatted_column_list_tmp
    ]
    return (
        table_data_df,
        formatted_column_list,
        column_end_index,
        table_name,
        verify_column_list,
    )


def verify_excel_sheet(format_name, sheet_name, verify_sheet_df, original_data_df):
    verify_sheet_df_tmp = verify_sheet_df[
        (verify_sheet_df["文件名"] == format_name)
        & (verify_sheet_df["sheet页"] == sheet_name)
    ]

    target_count = int(verify_sheet_df_tmp["数据量"].iloc[0])

    comparison_operator = verify_sheet_df["比较符"].iloc[0]

    original_count = original_data_df.shape[0]
    logger.info(f"original data count -> {original_count}")

    compare_result = compare_count(comparison_operator, original_count, target_count)
    return compare_result, original_count


def verify_excel_data(sheet_name, column_value_check, target_data):
    column_name_dict = {}
    column_name_list = []
    # Create an SQLiteManager instance
    sqlite = SQLite()

    # Data write into SQLite and check column content
    sqlite.to_sql(df=column_value_check, table_name="table1")

    field_values = column_value_check["字段名"]
    for item, field_value in field_values.items():
        tmp1 = sqlite.query(sql=f"SELECT * FROM table1 where  `字段名` ='{field_value}' ")
        column_name = field_value.replace(" ", "").replace("\t", "").replace("\n", "")
        logger.info(f"target_column -> {column_name}")
        # Get field value
        if column_name in target_data.columns:
            field_values_list = target_data[column_name].tolist()
            field_values_df = pd.DataFrame(field_values_list, columns=["field_value"])
            sqlite.to_sql(df=field_values_df, table_name="table2")

            # Non-empty validation
            logger.info(
                f"Non-empty validation -> sheet:'{sheet_name}' ,field_value:'{field_value}'"
            )
            if tmp1["非空校验"].iloc[0] == "1":
                tmp2 = sqlite.query(
                    sql=f"SELECT count(1) FROM table2 where IFNULL(field_value,'')='' "
                )

                query_result_count = tmp2.iloc[0, 0] if tmp2.shape[0] > 0 else 0
                if query_result_count == 0:
                    logger.info("Non-empty validation passed")
                else:
                    logger.info("Non-empty validation failed")
                    break
            else:
                logger.info("Non-empty validation is not needed")

            # Content format validation
            logger.info(
                f"Content format validation -> sheet:'{sheet_name}' ,field_value:'{field_value}'"
            )
            if tmp1["内容格式校验"].iloc[0] == "1" and tmp1["字段类型"].iloc[0] == "日期":
                result1 = True
                for value in field_values_list:
                    try:
                        datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        logger.info(f"Date format validation has failed -> {value}")
                        result1 = False
                        break

                if result1:
                    logger.info("Date format validation passed")
                else:
                    logger.info("Date format validation failed")
                    break

            elif tmp1["内容格式校验"].iloc[0] == "1" and tmp1["字段类型"].iloc[0] == "小数":
                result2 = True
                for value in field_values_list:
                    if isinstance(value, float):
                        logger.info(f"{value} is a decimal")
                    elif "." in str(value):
                        if (
                            str(float(value)) == str(value)
                            or float(str(value)) == value
                        ):
                            logger.info(f"value is a decimal -> {value}")
                        else:
                            logger.info(f"value is not a decimal -> {value}")
                            result2 = False
                    else:
                        logger.info(f"{value} is not a decimal")
                        result2 = False
                if result2:
                    logger.info("Decimal validation passed")
                else:
                    logger.info("Decimal validation failed")
                    break
            else:
                logger.info("Decimal validation is not needed")

            # Enumeration value validation
            logger.info(
                f"------------------------Enumeration value validation -> sheet:'{sheet_name}' ,field_value:'{field_value}'"
            )
            if tmp1["枚举值校验"].iloc[0] == "1":
                value = tmp1["枚举值校验规则"].iloc[0]
                value_list = value.split(",")  # 将字符串按逗号拆分为列表
                quoted_values = ", ".join(f"'{v.strip()}'" for v in value_list)
                logger.info(f"value -> {quoted_values}")

                tmp3 = sqlite.query(
                    sql=f"SELECT count(1) FROM table2 where  field_value  not in ({quoted_values}) and IFNULL(field_value,'')<>'' "
                )

                query_result_count = tmp3.iloc[0, 0] if tmp3.shape[0] > 0 else 0
                if query_result_count == 0:
                    logger.info(f"Enumeration value validation passed")
                else:
                    logger.info(f"Enumeration value validation failed")
                    continue
            else:
                logger.info("Enumeration value validation is not needed")

            # column_name_list = column_name_list.append(column_name)
            # field_content_validation = 0
            # column_name_dict[column_name] = field_content_validation
            logger.info(
                f"Validation passed -> sheet:'{sheet_name}' ,field_value:'{field_value}'"
            )
        else:
            continue
    # Close DB
    sqlite.close()

    # return column_name_list, column_name_dict


def get_excel_df(
    data_stream, header=0, nrows=None, custom_sheet_list=None, X=None, Y=None
):
    wb = load_workbook(data_stream, read_only=True, data_only=True)
    all_sheet_list = wb.sheetnames

    if custom_sheet_list is None:
        sheet_list = all_sheet_list
    else:
        sheet_list = custom_sheet_list
        for sheet in custom_sheet_list:
            if sheet not in all_sheet_list:
                raise ValueError(f"Sheet name '{sheet}' not found in the workbook.")

    wb.close()

    df_dict = {}

    for sheet_name in sheet_list:
        data_stream.seek(0)
        if X is None or X is None:
            usecols = None
        else:
            usecols = f"{X}:{Y}"
        df = pd.read_excel(
            data_stream,
            sheet_name=sheet_name,
            dtype=str,
            header=header,
            nrows=nrows,
            usecols=usecols,
            engine="openpyxl",
        )

        df_dict[sheet_name] = df

    logger.info(f"get_excel_df() upload table data")
    return df_dict, sheet_list


def transform_excel_column_to_index(column_name):
    index = 0
    for char in column_name:
        index *= 26
        index += ord(char) - ord("A") + 1
    return index
